from openpyxl import Workbook


def write_to_excel(data: dict, output_path="output.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    ws["A1"] = "Business Date"
    ws["B1"] = data.get("business_date")

    ws["A2"] = "Group Count"
    ws["B2"] = data.get("group_count")

    ws["A3"] = "Adult Count"
    ws["B3"] = data.get("adult_count")

    ws["A4"] = "Sales Value"
    ws["B4"] = data.get("sales_value")

    wb.save(output_path)
