import re
from openpyxl import Workbook


SAMPLE_TOP_TEXT = """
営業日 2026/04/05
組数 12組
大人人数 28人
"""

SAMPLE_SALES_TEXT = """
売上(税抜)
123,456
売上(税込)
135,801
"""


def extract_receipt_data(top_text: str, sales_text: str) -> dict:
    top_lines = [line.strip() for line in top_text.splitlines() if line.strip()]
    sales_lines = [line.strip() for line in sales_text.splitlines() if line.strip()]

    date_match = re.search(r"営業日\s*(\d{4}/\d{2}/\d{2})", top_text)
    business_date = date_match.group(1) if date_match else None

    group_count = None
    group_match = re.search(r"組数\s*(\d+)組", top_text)
    if group_match:
        group_count = int(group_match.group(1))
    else:
        group_match = re.search(r"組\s*(\d+)組", top_text)
        if group_match:
            group_count = int(group_match.group(1))

    adult_count = None

    match = re.search(r"大人人数\s*(\d+)人", top_text)
    if match:
        adult_count = int(match.group(1))

    if adult_count is None:
        match = re.search(r"大人男性\s*(\d+)人", top_text)
        if match:
            adult_count = int(match.group(1))

    if adult_count is None:
        for i, line in enumerate(top_lines):
            if re.search(r"組数?\s*\d+組", line):
                for j in range(i + 1, min(i + 4, len(top_lines))):
                    match = re.search(r"(\d+)人", top_lines[j])
                    if match:
                        adult_count = int(match.group(1))
                        break
                break

    sales_value = None
    in_tax_excluded_block = False

    for line in sales_lines:
        if "売上(税抜)" in line:
            in_tax_excluded_block = True
            continue

        if "売上(税込)" in line and in_tax_excluded_block:
            break

        if in_tax_excluded_block:
            candidates = re.findall(r"[\d][\d,\.]{2,}", line)
            if candidates:
                cleaned_candidates = [re.sub(r"[^\d]", "", c) for c in candidates]
                cleaned_candidates = [c for c in cleaned_candidates if len(c) >= 3]
                if cleaned_candidates:
                    sales_value = int(cleaned_candidates[0])
                    break

    if sales_value is None:
        candidates = re.findall(r"[\d][\d,\.]{2,}", sales_text)
        cleaned_candidates = [re.sub(r"[^\d]", "", c) for c in candidates]
        cleaned_candidates = [c for c in cleaned_candidates if len(c) >= 3]
        if cleaned_candidates:
            sales_value = int(cleaned_candidates[0])

    return {
        "business_date": business_date,
        "group_count": group_count,
        "adult_count": adult_count,
        "sales_value": sales_value,
    }


def write_to_excel(data: dict, output_path: str = "demo_output.xlsx") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    ws["A1"] = "Business Date"
    ws["B1"] = data.get("business_date")

    ws["A2"] = "Group Count"
    ws["B2"] = data.get("group_count")

    ws["A3"] = "Adult Count"
    ws["B3"] = data.get("adult_count")

    ws["A4"] = "Sales Value"
    ws["B4"] = data.get("sales_value")

    wb.save(output_path)


def main():
    data = extract_receipt_data(SAMPLE_TOP_TEXT, SAMPLE_SALES_TEXT)

    print("=== Parsed Result ===")
    print(f"Business Date: {data['business_date']}")
    print(f"Group Count  : {data['group_count']}")
    print(f"Adult Count  : {data['adult_count']}")
    print(f"Sales Value  : {data['sales_value']}")

    write_to_excel(data)
    print("\nSaved: demo_output.xlsx")


if __name__ == "__main__":
    main()
